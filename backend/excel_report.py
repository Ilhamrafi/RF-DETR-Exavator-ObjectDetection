# backend/excel_report.py

"""
Laporan Excel untuk hasil deteksi excavator/truck.

Fokus:
- Sheet "Tracking": urutan event per-ID dengan akumulasi ritase & passing.
- Sheet "Info Video": metadata video dan ringkasannya.
- Sheet "Detail Events": daftar event mentah (urut waktu).
- (Opsional) "Statistik Detail": ringkasan statistik tambahan bila tersedia.
"""

from __future__ import annotations

import os
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Union, Optional

import openpyxl
import supervision as sv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Logging
logger = logging.getLogger(__name__)


def generate_excel_report(
    output_path: str,
    passing_events: List[Dict[str, Union[float, int]]],
    ritase_events: List[Dict[str, Union[float, int]]],
    video_info: sv.VideoInfo,
    source_path: str,
) -> None:
    """
    Bangun laporan Excel berisi urutan event (passing/ritase) dan akumulasinya.

    Kolom "Tracking":
        - Nama File (ditulis sekali saat ID pertama muncul)
        - ID (bertambah saat ritase terdeteksi)
        - Datetime (timestamp event)
        - Ritase (akumulatif pada ID yang sama)
        - Passing (akumulatif pada ID yang sama)
        - Siklus (bertambah saat ritase terdeteksi)

    Args:
        output_path: Lokasi file Excel keluaran.
        passing_events: Daftar event passing dengan kunci: Frame, Detik, Confidence.
        ritase_events: Daftar event ritase dengan kunci: Frame, Detik, Confidence.
        video_info: Informasi video (width, height, fps, total_frames).
        source_path: Path video sumber (untuk nama file).
    """
    try:
        wb = openpyxl.Workbook()

        # Sheet utama
        tracking_sheet = wb.active
        tracking_sheet.title = "Tracking"

        # Header tabel Tracking
        headers = ["Nama File", "ID", "Datetime", "Ritase", "Passing", "Siklus"]
        for col, header in enumerate(headers, 1):
            cell = tracking_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        filename = os.path.basename(source_path)

        # Estimasi waktu mulai proses agar timestamp event relatif ke proses
        video_duration_seconds = video_info.total_frames / video_info.fps
        processing_start_time = datetime.now() - timedelta(seconds=video_duration_seconds)

        # Gabungkan seluruh event untuk diurutkan
        all_events: List[Dict[str, Union[str, int, float, datetime]]] = []

        # Event Passing
        for idx, event in enumerate(passing_events):
            frame = event["Frame"]
            seconds = event["Detik"]
            timestamp = processing_start_time + timedelta(seconds=seconds)
            all_events.append(
                {
                    "event_id": idx + 1,
                    "timestamp": timestamp,
                    "frame": frame,
                    "seconds": seconds,
                    "type": "passing",
                    "confidence": event["Confidence"],
                }
            )

        # Event Ritase
        for idx, event in enumerate(ritase_events):
            frame = event["Frame"]
            seconds = event["Detik"]
            timestamp = processing_start_time + timedelta(seconds=seconds)
            all_events.append(
                {
                    "event_id": idx + 1,
                    "timestamp": timestamp,
                    "frame": frame,
                    "seconds": seconds,
                    "type": "ritase",
                    "confidence": event["Confidence"],
                }
            )

        # Urutkan berdasarkan waktu
        all_events.sort(key=lambda x: x["timestamp"])  # type: ignore[index]

        # Penomoran ID & siklus (bertambah saat ritase muncul)
        current_id = 1
        current_siklus = 1
        cycle_counts = {1: {"ritase": 0, "passing": 0}}

        # Tandai setiap event dengan custom_id & siklus + catat akumulasi per-ID
        for event in all_events:
            event_type = event["type"]  # type: ignore[index]
            event["custom_id"] = current_id  # type: ignore[index]
            event["siklus"] = current_siklus  # type: ignore[index]

            if event_type == "passing":
                cycle_counts[current_id]["passing"] += 1
            elif event_type == "ritase":
                cycle_counts[current_id]["ritase"] += 1
                current_id += 1
                current_siklus += 1
                if current_id not in cycle_counts:
                    cycle_counts[current_id] = {"ritase": 0, "passing": 0}

        # Tulis ke sheet Tracking
        row = 2
        for event in all_events:
            custom_id = event["custom_id"]  # type: ignore[index]
            timestamp_str = event["timestamp"].strftime("%d-%m-%Y %H:%M:%S")  # type: ignore[index]
            siklus = event["siklus"]  # type: ignore[index]

            # Nama file hanya pada baris pertama untuk ID tersebut
            filename_value = ""
            if is_first_row_for_id(tracking_sheet, int(custom_id), row):
                filename_value = filename

            # Akumulasi hingga event saat ini (pada ID yang sama)
            passing_count = 0
            ritase_count = 0
            for prev_event in all_events:
                if prev_event["timestamp"] > event["timestamp"]:  # type: ignore[index]
                    continue
                if prev_event["custom_id"] != custom_id:  # type: ignore[index]
                    continue
                if prev_event["type"] == "passing":  # type: ignore[index]
                    passing_count += 1
                elif prev_event["type"] == "ritase":  # type: ignore[index]
                    ritase_count += 1

            tracking_sheet.cell(row=row, column=1, value=filename_value)
            tracking_sheet.cell(row=row, column=2, value=custom_id)
            tracking_sheet.cell(row=row, column=3, value=timestamp_str)
            tracking_sheet.cell(row=row, column=4, value=ritase_count)
            tracking_sheet.cell(row=row, column=5, value=passing_count)
            tracking_sheet.cell(row=row, column=6, value=siklus)
            row += 1

        # Lebar kolom Tracking
        column_widths = [30, 10, 20, 10, 10, 10]
        for i, width in enumerate(column_widths, 1):
            tracking_sheet.column_dimensions[get_column_letter(i)].width = width

        # Sheet Info Video
        info_sheet = wb.create_sheet("Info Video")
        info_sheet["A1"] = "Parameter"
        info_sheet["B1"] = "Nilai"
        info_sheet["A1"].font = Font(bold=True)
        info_sheet["B1"].font = Font(bold=True)

        durasi_detik = video_info.total_frames / float(video_info.fps)
        info_rows = [
            ("Nama File", os.path.basename(source_path)),
            ("Resolusi", f"{video_info.width}x{video_info.height}"),
            ("FPS", video_info.fps),
            ("Total Frame", video_info.total_frames),
            ("Durasi (detik)", durasi_detik),
            ("Durasi", f"{int(durasi_detik // 60)} menit {int(durasi_detik % 60)} detik"),
            ("Tanggal Pemrosesan", datetime.now().strftime("%d-%m-%Y %H:%M:%S")),
            ("Total Passing", len(passing_events)),
            ("Total Ritase", len(ritase_events)),
        ]

        for i, (param, value) in enumerate(info_rows, 2):
            info_sheet.cell(row=i, column=1, value=param)
            info_sheet.cell(row=i, column=2, value=value)

        info_sheet.column_dimensions["A"].width = 20
        info_sheet.column_dimensions["B"].width = 30

        # Sheet Detail Events
        detail_sheet = wb.create_sheet("Detail Events")

        detail_headers = ["No", "Tipe", "ID", "Siklus", "Frame", "Detik", "Timestamp", "Confidence"]
        for col, header in enumerate(detail_headers, 1):
            cell = detail_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        row = 2
        for event in all_events:
            detail_sheet.cell(row=row, column=1, value=event["event_id"])
            detail_sheet.cell(row=row, column=2, value=str(event["type"]).capitalize())  # type: ignore[index]
            detail_sheet.cell(row=row, column=3, value=event["custom_id"])  # type: ignore[index]
            detail_sheet.cell(row=row, column=4, value=event["siklus"])  # type: ignore[index]
            detail_sheet.cell(row=row, column=5, value=event["frame"])
            detail_sheet.cell(row=row, column=6, value=event["seconds"])
            detail_sheet.cell(row=row, column=7, value=event["timestamp"].strftime("%d-%m-%Y %H:%M:%S"))  # type: ignore[index]
            detail_sheet.cell(row=row, column=8, value=event["confidence"])
            row += 1

        # Lebar kolom Detail Events
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            detail_sheet.column_dimensions[col].width = 15
        detail_sheet.column_dimensions["G"].width = 20

        wb.save(output_path)
        logger.info("Laporan Excel berhasil disimpan: %s", output_path)

    except Exception as exc:
        logger.error("Gagal membuat laporan Excel: %s", str(exc))
        raise


def is_first_row_for_id(sheet: Worksheet, custom_id: int, current_row: int) -> bool:
    """
    Cek apakah baris saat ini adalah kemunculan pertama untuk `custom_id` pada sheet "Tracking".

    Args:
        sheet: Worksheet tujuan.
        custom_id: ID yang diperiksa.
        current_row: Baris yang sedang diisi.

    Returns:
        True jika ini baris pertama untuk ID tersebut; selain itu False.
    """
    for row in range(2, current_row):
        if sheet.cell(row=row, column=2).value == custom_id:
            return False
    return True


def generate_extended_report(
    output_path: str,
    passing_events: List[Dict[str, Union[float, int]]],
    ritase_events: List[Dict[str, Union[float, int]]],
    video_info: sv.VideoInfo,
    source_path: str,
    passing_stats: Optional[Dict] = None,
    ritase_stats: Optional[Dict] = None,
) -> None:
    """
    Buat laporan dasar lalu, bila tersedia, tambahkan sheet "Statistik Detail".

    Args:
        output_path: Lokasi file Excel keluaran.
        passing_events: Daftar event passing.
        ritase_events: Daftar event ritase.
        video_info: Informasi video sumber.
        source_path: Path video sumber.
        passing_stats: Statistik tambahan passing (opsional).
        ritase_stats: Statistik tambahan ritase (opsional).
    """
    # Bangun laporan dasar
    generate_excel_report(output_path, passing_events, ritase_events, video_info, source_path)

    # Jika tidak ada statistik, cukup laporan dasar saja
    if not passing_stats or not ritase_stats:
        return

    try:
        wb = openpyxl.load_workbook(output_path)
        detail_sheet = wb.create_sheet("Statistik Detail")

        # Judul
        detail_sheet["A1"] = "STATISTIK DETAIL DETEKSI"
        detail_sheet["A1"].font = Font(bold=True, size=14)
        detail_sheet.merge_cells("A1:D1")
        detail_sheet["A1"].alignment = Alignment(horizontal="center")

        # Blok Passing
        detail_sheet["A3"] = "Detail Passing:"
        detail_sheet["A3"].font = Font(bold=True)

        row = 4
        if passing_stats:
            for key, value in passing_stats.items():
                # Saring data yang besar agar tetap ringkas
                if key != "detail_excavator":
                    detail_sheet[f"A{row}"] = key.replace("_", " ").title()
                    detail_sheet[f"B{row}"] = str(value)
                    row += 1

        # Blok Ritase
        row += 1
        detail_sheet[f"A{row}"] = "Detail Ritase:"
        detail_sheet[f"A{row}"].font = Font(bold=True)
        row += 1

        if ritase_stats:
            for key, value in ritase_stats.items():
                # Saring kunci yang kurang relevan untuk ringkasan
                if key not in ["detail_truck", "current_cycle_has_ritase", "current_cycle_number"]:
                    detail_sheet[f"A{row}"] = key.replace("_", " ").title()
                    detail_sheet[f"B{row}"] = str(value)
                    row += 1

        # Lebar kolom
        detail_sheet.column_dimensions["A"].width = 30
        detail_sheet.column_dimensions["B"].width = 40

        wb.save(output_path)
        logger.info("Statistik detail berhasil ditambahkan ke laporan: %s", output_path)

    except Exception as exc:
        # Tidak raise ulang: laporan dasar sudah tersedia
        logger.error("Gagal menambahkan statistik detail: %s", str(exc))